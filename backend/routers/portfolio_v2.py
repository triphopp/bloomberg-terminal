"""
Portfolio v2 — Multi-account, multi-currency trade tracking.
Supports TH equity, US equity, and crypto accounts.
"""
import io
import json
import logging
import re
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from pydantic import BaseModel

from sources import market_data

from cache import TTLCache
from db import get_db
from portfolio_currency import (
    convert_amount,
    fx_rate as _fx,
    infer_instrument_currency,
    live_usd_thb,
    normalize_currency,
    realized_economic_pnl_in_report,
    realized_pnl_in_report,
    report_currency,
    trade_currency,
    trade_value_in_report,
)

# ── Price cache — TTL 60s per symbol, THB rate TTL 120s ──────────────────────
_price_cache: TTLCache = TTLCache(ttl=60,  maxsize=300)

router = APIRouter(prefix="/api/v2/portfolio")

logger = logging.getLogger("api.portfolio_v2")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _now() -> str:
    return datetime.utcnow().isoformat()


def _map_account_name(name: str) -> str:
    """Map spreadsheet account names to DB IDs."""
    n = str(name).lower().strip()
    if "finansia" in n:
        return "finansia"
    if "dime" in n:
        return "dime"
    if "innov" in n:
        return "innovestx"
    return "finansia"


def _to_float(v) -> Optional[float]:
    try:
        f = float(v)
        return f if f != 0 else None
    except (TypeError, ValueError):
        return None


def _to_float_or_zero(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _fmt_date(v) -> Optional[str]:
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s[:10] if len(s) >= 10 else s


def _get_yf_symbol(symbol: str, account_id: str) -> Optional[str]:
    """Convert trade symbol to yfinance-compatible format."""
    if not symbol:
        return None
    sym = symbol.strip().upper()
    # Skip options (PUT_INTC, CALL_INTC etc.)
    if sym.startswith(("PUT_", "CALL_")):
        return None
    if account_id == "finansia":
        return f"{sym}.BK"
    if account_id == "innovestx":
        # Already yf-compatible: BTC-USD, BTC-THB, SOL-USD etc.
        if "-" in sym:
            return sym
        # BTCTHB → BTC-THB, SOLTHB → SOL-THB
        if sym.endswith("THB") and len(sym) > 3:
            base = sym[:-3]
            return f"{base}-THB"
        return None
    return sym  # Dime: US symbols as-is


def _trade_yf_symbol(row: dict) -> Optional[str]:
    """Provider symbol for a trade row.

    Prefer the resolved_symbol persisted at entry time (symbol resolver,
    plans/port-redesign.md); fall back to the legacy account-id heuristic
    for rows created before the resolver existed (F06 transition path).
    """
    rs = row.get("resolved_symbol")
    if rs:
        return str(rs).strip().upper()
    return _get_yf_symbol(row.get("symbol") or "", row.get("account_id") or "")


def _position_currency(row: dict) -> str:
    """Compatibility alias; stored currency wins, inference is NULL-only."""
    return trade_currency(row)


def _conv(amount: float, ccy: str, base: str, thb_per_usd: float) -> float:
    """Compatibility wrapper for live THB/USD conversions."""
    return convert_amount(amount, ccy, base, stored_thb_rate=thb_per_usd)


def _get_live_price(symbol: str, account_id: str) -> Optional[float]:
    yf_sym = _get_yf_symbol(symbol, account_id)
    if not yf_sym:
        return None
    cached = _price_cache.get(yf_sym)
    if cached is not None:
        return cached
    try:
        info = market_data.get_fast_info(yf_sym)
        price = getattr(info, "last_price", None) or getattr(info, "regular_market_price", None)
        result = float(price) if price else None
        if result:
            _price_cache.set(yf_sym, result)
        return result
    except Exception:
        return None


_prev_cache: TTLCache = TTLCache(ttl=300, maxsize=300)

# Stale-while-revalidate layer: last known good quote per symbol (never expires).
# When the 60s TTL cache misses but a stale value exists, we serve the stale
# value immediately and refresh in a background thread — so page loads never
# block on yfinance after the very first fetch of a symbol.
_stale_quotes: dict[str, dict] = {}
_stale_lock = threading.Lock()
_refresh_inflight: set[str] = set()


def _fetch_one_quote(sym: str) -> dict:
    """Fetch a single symbol's quote via fast_info. Returns {"price","prev_close"}."""
    try:
        info = market_data.get_fast_info(sym)
        price = getattr(info, "last_price", None) or getattr(info, "regular_market_price", None)
        prev = getattr(info, "previous_close", None)
        return {"price": float(price) if price else None,
                "prev_close": float(prev) if prev else None}
    except Exception:
        return {"price": None, "prev_close": None}


def _store_quote(sym: str, price, prev) -> None:
    if price:
        _price_cache.set(sym, price)
    if prev:
        _prev_cache.set(sym, prev)
    if price or prev:
        with _stale_lock:
            _stale_quotes[sym] = {"price": price, "prev_close": prev}


def _fetch_symbols_now(to_fetch: list[str]) -> dict[str, dict]:
    """Blocking fetch: batch download, then parallel per-symbol fallback."""
    result: dict[str, dict] = {}
    try:
        batch = market_data.download_quotes(to_fetch)
        for sym in to_fetch:
            snap = batch.quotes.get(sym) if hasattr(batch, "quotes") else None
            price = snap.last_price if snap else None
            prev = snap.previous_close if snap else None
            result[sym] = {"price": price, "prev_close": prev}
            _store_quote(sym, price, prev)
    except Exception:
        pass

    # Fill missing with individual fallback — parallel, not one-at-a-time
    missing = [s for s in to_fetch if result.get(s, {}).get("price") is None]
    if missing:
        with ThreadPoolExecutor(max_workers=min(8, len(missing))) as ex:
            for sym, q in zip(missing, ex.map(_fetch_one_quote, missing)):
                result[sym] = q
                _store_quote(sym, q["price"], q["prev_close"])
    return result


def _refresh_in_background(symbols: list[str]) -> None:
    """Kick off a daemon thread to refresh quotes; coalesces duplicate requests."""
    with _stale_lock:
        todo = [s for s in symbols if s not in _refresh_inflight]
        _refresh_inflight.update(todo)
    if not todo:
        return

    def _run():
        try:
            _fetch_symbols_now(todo)
        finally:
            with _stale_lock:
                _refresh_inflight.difference_update(todo)

    threading.Thread(target=_run, daemon=True).start()


def _batch_fetch_prices(symbols: list[str]) -> dict[str, dict]:
    """Batch-fetch current prices + prev_close for multiple yfinance symbols.
    Returns dict[sym] = {"price": float|None, "prev_close": float|None}.
    Fresh-cache hit → served as-is. TTL-expired but stale value known → stale
    served instantly + background refresh. Never-seen symbols → blocking fetch.
    """
    result: dict[str, dict] = {}
    cold: list[str] = []      # never seen — must block
    stale_syms: list[str] = []  # expired but have last-known value

    for sym in symbols:
        cached_price = _price_cache.get(sym)
        if cached_price is not None:
            result[sym] = {"price": cached_price, "prev_close": _prev_cache.get(sym)}
            continue
        with _stale_lock:
            stale = _stale_quotes.get(sym)
        if stale is not None:
            result[sym] = dict(stale)
            stale_syms.append(sym)
        else:
            cold.append(sym)

    if stale_syms:
        _refresh_in_background(stale_syms)
    if cold:
        result.update(_fetch_symbols_now(cold))

    return result


def _get_thb_per_usd() -> float:
    return live_usd_thb()


def _capture_thb_rate(
    currency: str,
    date: Optional[str],
    provided: Optional[float] = None,
    *,
    conn=None,
) -> float:
    """THB per one native-currency unit for a transaction date."""
    if currency == "THB":
        return 1.0
    try:
        explicit = float(provided or 0)
    except (TypeError, ValueError):
        explicit = 0.0
    if explicit > 1:
        return explicit
    return _fx(currency, "THB", date=date, conn=conn)


# ── Audit helpers ────────────────────────────────────────────────────────────

def _write_audit_log(
    conn,
    trade_id: str,
    action: str,
    old_row: dict,
    new_values: dict | None = None,
    reason: str = "",
) -> None:
    """Append one immutable event to trade_audit_log."""
    fields_changed: dict = {}
    if new_values:
        for k, v in new_values.items():
            old_v = old_row.get(k)
            if old_v != v:
                fields_changed[k] = {"old": old_v, "new": v}

    conn.execute(
        """INSERT INTO trade_audit_log
               (trade_id, action, fields_changed, reason, snapshot)
           VALUES (?, ?, ?, ?, ?)""",
        (
            trade_id,
            action,
            json.dumps(fields_changed),
            reason or "",
            json.dumps({k: old_row.get(k) for k in [
                "symbol", "price_entry", "volume", "date_entry",
                "price_exit", "date_exit", "win_loss", "pnl_amount",
                "price_stoploss", "price_target", "note",
            ]}),
        ),
    )


# ── Pydantic Models ──────────────────────────────────────────────────────────

class AccountIn(BaseModel):
    id: str
    name: str
    broker: str = ""
    country: str = "TH"
    currency: str = "THB"
    account_type: str = "equity"


class AccountPatch(BaseModel):
    name: Optional[str] = None
    broker: Optional[str] = None
    is_active: Optional[int] = None


class TradeIn(BaseModel):
    account_id: str
    symbol: str
    resolved_symbol: Optional[str] = None
    market: Optional[str] = None
    sector: str = ""
    date_entry: str
    date_exit: Optional[str] = None
    price_entry: float = 0
    price_exit: Optional[float] = None
    price_stoploss: Optional[float] = None
    price_target: Optional[float] = None
    volume: float = 0
    amount: Optional[float] = None
    pnl_amount: Optional[float] = None
    win_loss: str = "P"
    pnl_percent: Optional[float] = None
    currency: Optional[str] = None
    exchange_rate: Optional[float] = None
    exit_exchange_rate: Optional[float] = None
    strategy_name: str = ""
    entry_trigger: str = ""
    exit_trigger: str = ""
    market_trend: str = ""
    news_sentiment: str = ""
    expectation_based: str = ""
    factor_based: str = ""
    fear_greed_index: str = ""
    vix_index: str = ""
    note: str = ""


class TradePatch(BaseModel):
    # Entry fields (แก้ไขข้อมูลที่กรอกผิด)
    symbol:          Optional[str]   = None
    sector:          Optional[str]   = None
    date_entry:      Optional[str]   = None
    price_entry:     Optional[float] = None
    price_stoploss:  Optional[float] = None
    price_target:    Optional[float] = None
    volume:          Optional[float] = None
    strategy_name:   Optional[str]   = None
    entry_trigger:   Optional[str]   = None
    market_trend:    Optional[str]   = None
    note:            Optional[str]   = None
    # Exit fields
    date_exit:       Optional[str]   = None
    price_exit:      Optional[float] = None
    pnl_amount:      Optional[float] = None
    win_loss:        Optional[str]   = None
    pnl_percent:     Optional[float] = None
    exit_trigger:    Optional[str]   = None
    # Audit meta — NOT persisted to trades table, only to audit log
    adjustment_reason: Optional[str] = None


class CashIn(BaseModel):
    account_id: str
    date: str
    income: float = 0
    investment: float = 0
    exchange_rate: float = 1
    note: str = ""


class CashTransferIn(BaseModel):
    from_account_id: str
    to_account_id: str
    date: str
    amount: float
    note: str = ""


class DividendIn(BaseModel):
    account_id: str
    asset: str
    ex_date: Optional[str] = None
    pay_date: Optional[str] = None
    amount_per_unit: float = 0
    total_received: float = 0
    reinvested_amount: float = 0
    reinvest_asset: str = ""
    reinvest_price: float = 0
    reinvest_units: float = 0
    currency: Optional[str] = None


# ── Symbol Resolver (plans/port-redesign.md Step 1) ──────────────────────────
# Resolve once at write time: user types a bare ticker, we search only the
# markets the account trades and return canonical provider symbols.  The
# chosen resolved_symbol is persisted on the trade so no read path ever has
# to guess exchange suffixes again (root cause of F06).

_MARKET_SUFFIX = {"TH": ".BK", "US": ""}
_MARKET_CURRENCY = {"TH": "THB", "US": "USD"}
_resolve_cache: TTLCache = TTLCache(ttl=3600, maxsize=500)


def _classify_market(sym: str, quote_type: str = "") -> Optional[str]:
    s = sym.upper()
    qt = (quote_type or "").upper()
    # Forex (BHDHKD=X) and index (^GSPC) tickers are not tradeable equities —
    # they otherwise slip through the "no dot ⇒ US" rule and pollute matches.
    if "=" in s or s.startswith("^"):
        return None
    if qt in ("CURRENCY", "INDEX", "FUTURE"):
        return None
    if s.endswith(".BK"):
        return "TH"
    if "CRYPTO" in qt or ("-" in s and s.rsplit("-", 1)[-1] in ("USD", "THB", "USDT")):
        return "CRYPTO"
    if "." not in s:
        return "US"
    return None  # other exchange suffixes — unsupported for now


def _account_markets(acc: dict) -> list[str]:
    raw = acc.get("markets")
    if raw:
        try:
            parsed = [str(m).upper() for m in json.loads(raw)]
            if parsed:
                return parsed
        except Exception:
            pass
    return ["CRYPTO"] if acc.get("account_type") == "crypto" else ["US", "TH"]


@router.get("/resolve-symbol")
def resolve_symbol(q: str = Query(..., min_length=1), account_id: str = Query(...)):
    query = q.strip().upper()
    if not query:
        raise HTTPException(status_code=400, detail="Empty query")
    with get_db() as conn:
        row = conn.execute("SELECT * FROM portfolio_accounts WHERE id = ?",
                           (account_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Unknown account")
    acc = dict(row)
    markets = _account_markets(acc)
    home = str(acc.get("country") or "").upper()

    # Home country must be part of the key: two accounts can share the same
    # markets list (e.g. ["US","TH"]) but rank differently by home market, so
    # omitting `home` let one account serve another's ordering — a TH account
    # would then surface US matches first (F06-style cross-market mix-up).
    cache_key = f"{query}|{home}|{','.join(markets)}"
    cached = _resolve_cache.get(cache_key)
    if cached is not None:
        return cached

    matches: list[dict] = []
    seen: set[str] = set()

    def _add(sym: str, market: str, name: str, exchange: str) -> None:
        if sym in seen or market not in markets:
            return
        seen.add(sym)
        matches.append({
            "resolved_symbol": sym,
            "market": market,
            "currency": _MARKET_CURRENCY.get(market)
                        or (sym.rsplit("-", 1)[-1] if "-" in sym else None),
            "name": name,
            "exchange": exchange,
        })

    def _add_from_result(r, market_hint: Optional[str] = None) -> None:
        sym = str(getattr(r, "symbol", "") or "").upper()
        if not sym:
            return
        market = market_hint or _classify_market(sym, str(getattr(r, "quote_type", "") or ""))
        if not market:
            return
        name = str(getattr(r, "long_name", "") or getattr(r, "short_name", "") or "")
        _add(sym, market, name, str(getattr(r, "exchange", "") or ""))

    # 1) free-text search (catches cross-market listings).  Keep only results
    # whose ticker starts with the query — this is a ticker field, and Yahoo's
    # name-relevance matches (e.g. "TU" → APPS "Digital Turbine") are noise.
    try:
        for r in market_data.search(query, max_results=10):
            sym = str(getattr(r, "symbol", "") or "").upper()
            if sym.split(".")[0].split("-")[0].startswith(query):
                _add_from_result(r)
    except Exception:
        logger.warning("resolve-symbol search failed for %r", query, exc_info=True)

    # 2) direct suffix probes — exact tickers the free-text search often misses
    probes: list[tuple[str, str]] = []  # (candidate, market)
    for m in markets:
        if m == "CRYPTO":
            if "-" in query:
                probes.append((query, m))
            elif query.endswith(("THB", "USD")) and len(query) > 3:
                base = query[:-3]
                # THB pairs are mostly delisted on Yahoo — offer USD too
                probes.append((f"{base}-{query[-3:]}", m))
                probes.append((f"{base}-USD", m))
            else:
                probes.append((f"{query}-USD", m))
        else:
            suffix = _MARKET_SUFFIX.get(m)
            if suffix is not None:
                probes.append((f"{query}{suffix}", m))
    for cand, m in probes:
        if cand in seen:
            continue
        found = False
        try:
            for r in market_data.search(cand, max_results=3):
                if str(getattr(r, "symbol", "") or "").upper() == cand:
                    _add_from_result(r, market_hint=m)
                    found = True
                    break
        except Exception:
            pass
        if not found:
            # Yahoo Search doesn't index some symbols (e.g. BTC-THB) that the
            # quote API prices fine — validate via a live quote instead.
            try:
                snap = market_data.get_fast_info(cand)
                if snap is not None and snap.last_price is not None:
                    _add(cand, m, "", "")
            except Exception:
                continue

    # rank: account home country first, then declared markets order
    order = {mk: i for i, mk in enumerate(markets)}
    matches.sort(key=lambda x: (x["market"] != home, order.get(x["market"], 99)))
    result = {"query": query, "markets": markets, "matches": matches}
    if matches:  # don't cache misses — provider hiccup would stick for an hour
        _resolve_cache.set(cache_key, result)
    return result


# ── Accounts ─────────────────────────────────────────────────────────────────

@router.get("/accounts")
def list_accounts():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM portfolio_accounts WHERE is_active = 1 ORDER BY created_at"
        ).fetchall()
    return [dict(r) for r in rows]


@router.post("/accounts", status_code=201)
def create_account(body: AccountIn):
    with get_db() as conn:
        conn.execute("""
            INSERT INTO portfolio_accounts (id, name, broker, country, currency, account_type)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (body.id, body.name, body.broker, body.country, body.currency, body.account_type))
    return {"ok": True, "id": body.id}


@router.patch("/accounts/{account_id}")
def patch_account(account_id: str, body: AccountPatch):
    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    with get_db() as conn:
        conn.execute(f"UPDATE portfolio_accounts SET {set_clause} WHERE id = ?",
                     list(updates.values()) + [account_id])
    return {"ok": True}


@router.delete("/accounts/{account_id}")
def delete_account(account_id: str):
    with get_db() as conn:
        n_trades = conn.execute(
            "SELECT COUNT(*) FROM trades WHERE account_id = ?", (account_id,)
        ).fetchone()[0]
        if n_trades > 0:
            raise HTTPException(
                status_code=409,
                detail=f"Account has {n_trades} trade(s). Delete or move them first.",
            )
        cur = conn.execute("DELETE FROM portfolio_accounts WHERE id = ?", (account_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Account not found")
        conn.execute("DELETE FROM cash_ledger WHERE account_id = ?", (account_id,))
        conn.execute("DELETE FROM dividends   WHERE account_id = ?", (account_id,))
    return {"ok": True}


# ── Trades ───────────────────────────────────────────────────────────────────

@router.get("/trades")
def list_trades(
    account_id: Optional[str] = Query(None),
    win_loss: Optional[str] = Query(None),
    symbol: Optional[str] = Query(None),
    base_currency: Optional[str] = Query(None),
    limit: int = Query(500),
    offset: int = Query(0),
):
    where, params = [], []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)
    if win_loss:
        where.append("win_loss = ?")
        params.append(win_loss.upper())
    if symbol:
        where.append("symbol LIKE ?")
        params.append(f"%{symbol.upper()}%")
    sql = "SELECT * FROM trades"
    if where:
        sql += " WHERE " + " AND ".join(where)
    # Order by most-recent activity: a freshly-closed trade keeps its old
    # date_entry, so sort by exit date when present (else entry) to surface
    # recent sells at the top instead of burying them by entry date.
    sql += " ORDER BY COALESCE(date_exit, date_entry) DESC, created_at DESC LIMIT ? OFFSET ?"
    params += [limit, offset]
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
        trades = [dict(r) for r in rows]
        if base_currency:
            base = report_currency(base_currency)
            for trade in trades:
                amount = abs(_to_float_or_zero(trade.get("amount"))) or (
                    _to_float_or_zero(trade.get("price_entry"))
                    * _to_float_or_zero(trade.get("volume"))
                )
                trade["amount_base"] = round(
                    trade_value_in_report(trade, amount, base, when="entry", conn=conn), 2
                )
                trade["price_entry_base"] = round(
                    trade_value_in_report(
                        trade,
                        _to_float_or_zero(trade.get("price_entry")),
                        base,
                        when="entry",
                        conn=conn,
                    ),
                    6,
                )
                if trade.get("price_exit") is not None:
                    trade["price_exit_base"] = round(
                        trade_value_in_report(
                            trade,
                            _to_float_or_zero(trade.get("price_exit")),
                            base,
                            when="exit",
                            conn=conn,
                        ),
                        6,
                    )
                if trade.get("date_exit") or trade.get("win_loss") in {"W", "L"}:
                    trade["pnl_base"] = round(
                        realized_pnl_in_report(trade, base, conn=conn), 2
                    )
    return {"trades": trades, "thb_per_usd": _get_thb_per_usd()}


@router.post("/trades", status_code=201)
def create_trade(body: TradeIn):
    trade_id = str(uuid.uuid4())
    with get_db() as conn:
        acc = conn.execute("SELECT currency FROM portfolio_accounts WHERE id = ?",
                           (body.account_id,)).fetchone()
        if not acc:
            raise HTTPException(status_code=404, detail="Unknown account")
        # Deterministic resolver evidence wins. Explicit currency supports the
        # confirmed manual-override path; account currency is only the last fallback.
        inferred = infer_instrument_currency(
            body.market, body.resolved_symbol, body.symbol, None
        )
        currency = (
            inferred
            or normalize_currency(body.currency)
            or normalize_currency(acc["currency"], "THB")
            or "THB"
        )
        entry_fx = _capture_thb_rate(
            currency, body.date_entry, body.exchange_rate, conn=conn
        )
        exit_fx = None
        if body.date_exit or body.win_loss.upper() != "P":
            exit_fx = _capture_thb_rate(
                currency,
                body.date_exit or body.date_entry,
                body.exit_exchange_rate,
                conn=conn,
            )
        conn.execute("""
            INSERT INTO trades (id, account_id, symbol, resolved_symbol, market,
                sector, date_entry, date_exit,
                price_entry, price_exit, price_stoploss, price_target, volume, amount,
                pnl_amount, win_loss, pnl_percent, currency, exchange_rate, exit_exchange_rate,
                strategy_name, entry_trigger, exit_trigger, market_trend,
                news_sentiment, expectation_based, factor_based,
                fear_greed_index, vix_index, note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (trade_id, body.account_id, body.symbol.upper(),
              (body.resolved_symbol or "").upper() or None,
              (body.market or "").upper() or None, body.sector,
              body.date_entry, body.date_exit,
              body.price_entry, body.price_exit, body.price_stoploss, body.price_target,
              body.volume, body.amount, body.pnl_amount, body.win_loss.upper(),
              body.pnl_percent, currency, entry_fx, exit_fx,
              body.strategy_name, body.entry_trigger, body.exit_trigger,
              body.market_trend, body.news_sentiment, body.expectation_based,
              body.factor_based, body.fear_greed_index, body.vix_index, body.note))
    return {
        "ok": True,
        "id": trade_id,
        "currency": currency,
        "exchange_rate": entry_fx,
        "exit_exchange_rate": exit_fx,
    }


@router.patch("/trades/{trade_id}")
def patch_trade(trade_id: str, body: TradePatch):
    updates = body.model_dump(exclude_none=True)
    reason = updates.pop("adjustment_reason", "") or ""
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")
    if "symbol" in updates and updates["symbol"]:
        updates["symbol"] = updates["symbol"].upper()
    if "win_loss" in updates and updates["win_loss"]:
        updates["win_loss"] = updates["win_loss"].upper()
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    with get_db() as conn:
        old = conn.execute("SELECT * FROM trades WHERE id = ?", (trade_id,)).fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Trade not found")
        old_dict = dict(old)
        cur = conn.execute(f"UPDATE trades SET {set_clause} WHERE id = ?",
                           list(updates.values()) + [trade_id])
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Trade not found")
        _write_audit_log(conn, trade_id, "PATCH", old_dict, updates, reason)
    return {"ok": True}


@router.delete("/trades/{trade_id}")
def delete_trade(trade_id: str):
    with get_db() as conn:
        old = conn.execute("SELECT * FROM trades WHERE id = ?", (trade_id,)).fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Trade not found")
        _write_audit_log(conn, trade_id, "DELETE", dict(old), None, "")
        conn.execute("DELETE FROM trades WHERE id = ?", (trade_id,))
    return {"ok": True}


# ── Trade Audit Log ─────────────────────────────────────────────────────────

@router.get("/trades/{trade_id}/audit-log")
def get_trade_audit_log(trade_id: str, limit: int = Query(50)):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM trade_audit_log WHERE trade_id = ? ORDER BY created_at DESC LIMIT ?",
            (trade_id, limit),
        ).fetchall()
    entries = []
    for r in rows:
        d = dict(r)
        try:
            d["fields_changed"] = json.loads(d["fields_changed"] or "{}")
        except Exception:
            d["fields_changed"] = {}
        try:
            d["snapshot"] = json.loads(d["snapshot"] or "{}")
        except Exception:
            d["snapshot"] = {}
        entries.append(d)
    return {"audit_log": entries}


@router.get("/audit-log")
def get_recent_audit_log(limit: int = Query(100), account_id: Optional[str] = Query(None)):
    with get_db() as conn:
        if account_id and account_id != "all":
            rows = conn.execute(
                """SELECT l.* FROM trade_audit_log l
                   JOIN trades t ON l.trade_id = t.id
                   WHERE t.account_id = ?
                   ORDER BY l.created_at DESC LIMIT ?""",
                (account_id, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM trade_audit_log ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
    entries = []
    for r in rows:
        d = dict(r)
        try:
            d["fields_changed"] = json.loads(d["fields_changed"] or "{}")
        except Exception:
            d["fields_changed"] = {}
        try:
            d["snapshot"] = json.loads(d["snapshot"] or "{}")
        except Exception:
            d["snapshot"] = {}
        entries.append(d)
    return {"audit_log": entries}


# ── Position Cost Overrides (manual avg cost correction) ────────────────────

class CostOverrideIn(BaseModel):
    account_id: str
    symbol: str
    avg_cost: float
    reason: Optional[str] = ""

@router.get("/cost-overrides")
def get_cost_overrides(account_id: Optional[str] = Query(None)):
    with get_db() as conn:
        if account_id and account_id != "all":
            rows = conn.execute(
                "SELECT * FROM position_cost_overrides WHERE account_id = ?",
                (account_id,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM position_cost_overrides").fetchall()
    return [dict(r) for r in rows]

@router.post("/cost-overrides")
def set_cost_override(body: CostOverrideIn):
    with get_db() as conn:
        conn.execute(
            """INSERT INTO position_cost_overrides (account_id, symbol, avg_cost, reason, updated_at)
               VALUES (?, ?, ?, ?, datetime('now'))
               ON CONFLICT(account_id, symbol) DO UPDATE SET
               avg_cost = excluded.avg_cost, reason = excluded.reason, updated_at = excluded.updated_at""",
            (body.account_id, body.symbol, body.avg_cost, body.reason or "")
        )
    return {"ok": True}

@router.delete("/cost-overrides/{account_id}/{symbol}")
def delete_cost_override(account_id: str, symbol: str):
    with get_db() as conn:
        conn.execute(
            "DELETE FROM position_cost_overrides WHERE account_id = ? AND symbol = ?",
            (account_id, symbol)
        )
    return {"ok": True}


# ── Sell / Partial Sell ──────────────────────────────────────────────────────

# ── Bulk Sector Re-normalize (one-time migration tool) ─────────────────────

@router.post("/trades/bulk-patch-sector")
def bulk_patch_sector(body: dict):
    """
    body = {"mappings": {"GOLD": "Materials", ...}, "account_id": "dime"}  # optional filter
    """
    import os
    if os.getenv("ALLOW_DANGEROUS_OPS", "") != "1":
        raise HTTPException(status_code=403, detail="Set ALLOW_DANGEROUS_OPS=1 env var to enable")

    mappings: dict = body.get("mappings", {})
    account_id: str | None = body.get("account_id")
    if not mappings:
        raise HTTPException(status_code=400, detail="No mappings provided")
    updated = 0
    with get_db() as conn:
        for old_val, new_val in mappings.items():
            where = "sector = ?"
            params: list = [old_val]
            if account_id:
                where += " AND account_id = ?"
                params.append(account_id)
            cur = conn.execute(f"UPDATE trades SET sector = ? WHERE {where}",
                               [new_val] + params)
            updated += cur.rowcount
    return {"ok": True, "updated": updated}


class SellIn(BaseModel):
    trade_id: str                        # ID of the open position being sold
    sell_volume: float = 0               # How many shares/units to sell (0 = all)
    sell_price: float = 0                # Exit price
    sell_date: str                       # Date of sale (YYYY-MM-DD)
    commission: float = 0                # Optional commission


@router.post("/sell", status_code=201)
def sell_position(body: SellIn):
    """Close or partially reduce an open position.

    Full sell (sell_volume = 0 or >= position volume):
        Updates the trade: sets date_exit, price_exit, pnl_amount, win_loss

    Partial sell (sell_volume < position volume):
        1. Creates a new CLOSED trade for the sold portion
        2. Reduces the original trade's volume (remaining shares stay open)
    """
    with get_db() as conn:
        # Find the open position
        row = conn.execute(
            "SELECT * FROM trades WHERE id = ? AND win_loss = 'P'",
            (body.trade_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Open position not found")

        pos = dict(row)
        total_volume = float(pos["volume"])
        sell_vol = float(body.sell_volume) if body.sell_volume > 0 else total_volume

        if sell_vol > total_volume:
            raise HTTPException(status_code=400, detail=f"Sell volume ({sell_vol}) exceeds position volume ({total_volume})")

        # Cost override takes priority; fall back to AVCO across all open lots
        override_row = conn.execute(
            "SELECT avg_cost FROM position_cost_overrides WHERE account_id = ? AND symbol = ?",
            (pos["account_id"], pos["symbol"])
        ).fetchone()
        if override_row:
            avg_cost = float(override_row["avg_cost"])
        else:
            all_lots = conn.execute(
                "SELECT price_entry, volume FROM trades WHERE account_id = ? AND symbol = ? AND win_loss = 'P'",
                (pos["account_id"], pos["symbol"])
            ).fetchall()
            total_vol_all = sum(float(l["volume"]) for l in all_lots)
            if total_vol_all > 0:
                avg_cost = sum(float(l["price_entry"]) * float(l["volume"]) for l in all_lots) / total_vol_all
            else:
                avg_cost = float(pos["price_entry"])

        entry_price = avg_cost
        exit_price  = float(body.sell_price)
        remaining   = round(total_volume - sell_vol, 8)
        pos_ccy = _position_currency(pos)
        exit_fx = _capture_thb_rate(pos_ccy, body.sell_date, conn=conn)

        if remaining <= 1e-8:
            # ── Full sell: close the position ────────────────────────────────
            pnl = round((exit_price - entry_price) * total_volume, 2)
            pnl_pct = round(((exit_price / entry_price) - 1) * 100, 2) if entry_price > 0 else 0
            pnl_net = round(pnl - float(body.commission), 2)
            wl = "W" if pnl_net >= 0 else "L"

            new_vals = {
                "date_exit": body.sell_date, "price_exit": exit_price,
                "pnl_amount": pnl_net, "win_loss": wl, "pnl_percent": pnl_pct,
                "exit_exchange_rate": exit_fx,
            }
            conn.execute(
                """UPDATE trades SET date_exit = ?, price_exit = ?,
                   pnl_amount = ?, win_loss = ?, pnl_percent = ?,
                   exit_exchange_rate = ?, note = note || ?
                   WHERE id = ?""",
                (body.sell_date, exit_price, pnl_net, wl, pnl_pct, exit_fx,
                  f"\n[SOLD {body.sell_date}] @ {exit_price} | P&L: {pnl_net}",
                  body.trade_id),
            )
            _write_audit_log(conn, body.trade_id, "SELL_FULL", pos, new_vals,
                             f"full sell {total_volume} @ {exit_price}")

            return {
                "ok": True,
                "action": "full_sell",
                "trade_id": body.trade_id,
                "pnl_amount": pnl_net,
                "pnl_percent": pnl_pct,
                "win_loss": wl,
            }
        else:
            # ── Partial sell: split into sold + remaining ────────────────────
            sold_volume = sell_vol
            sold_pnl = round((exit_price - entry_price) * sold_volume, 2)
            sold_pnl_pct = round(((exit_price / entry_price) - 1) * 100, 2) if entry_price > 0 else 0
            sold_pnl_net = round(sold_pnl - float(body.commission), 2)
            sold_wl = "W" if sold_pnl_net >= 0 else "L"

            import uuid as _uuid
            sold_id = str(_uuid.uuid4())
            conn.execute(
                """INSERT INTO trades (id, account_id, symbol, resolved_symbol, market,
                   sector, date_entry, date_exit,
                   price_entry, price_exit, volume, pnl_amount, win_loss, pnl_percent,
                   currency, exchange_rate, exit_exchange_rate, strategy_name, note)
                   SELECT ?, account_id, symbol, resolved_symbol, market,
                   sector, date_entry, ?,
                   ?, ?, ?, ?, ?, ?,
                   currency, exchange_rate, ?, strategy_name, ?
                   FROM trades WHERE id = ?""",
                (sold_id, body.sell_date, avg_cost, exit_price, sold_volume,
                  sold_pnl_net, sold_wl, sold_pnl_pct, exit_fx,
                  "",
                  body.trade_id),
            )

            # Reduce the remaining open lot. The partial sell is captured in the
            # audit log (below) — no auto-note appended to keep notes user-owned.
            conn.execute(
                "UPDATE trades SET volume = ? WHERE id = ?",
                (remaining, body.trade_id),
            )

            # Log on BOTH the original lot (volume reduced) and the new sold record
            _write_audit_log(conn, body.trade_id, "SELL_PARTIAL", pos,
                             {"volume": remaining},
                             f"partial sell {sold_volume} @ {exit_price}, avg_cost={round(avg_cost,4)}, remaining={remaining}")
            sold_snap = {k: pos.get(k) for k in ["symbol", "price_entry", "date_entry"]}
            _write_audit_log(conn, sold_id, "SELL_PARTIAL_CREATED", sold_snap,
                             {"volume": sold_volume, "price_entry": avg_cost,
                               "price_exit": exit_price,
                               "win_loss": sold_wl, "pnl_amount": sold_pnl_net,
                               "exit_exchange_rate": exit_fx},
                             f"created from partial sell of {body.trade_id}")

            return {
                "ok": True,
                "action": "partial_sell",
                "sold_trade_id": sold_id,
                "remaining_trade_id": body.trade_id,
                "sold_volume": sold_volume,
                "remaining_volume": remaining,
                "pnl_amount": sold_pnl_net,
                "pnl_percent": sold_pnl_pct,
                "win_loss": sold_wl,
            }


class SellAllLotsIn(BaseModel):
    account_id: str
    symbol: str
    sell_price: float
    sell_date: str
    commission: float = 0


@router.post("/sell-all-lots", status_code=201)
def sell_all_lots(body: SellAllLotsIn):
    """Close ALL open lots for a given account+symbol in one call.
    Uses AVCO across all lots. Calls the same full-sell logic per lot.
    """
    with get_db() as conn:
        lots = conn.execute(
            "SELECT * FROM trades WHERE account_id = ? AND symbol = ? AND win_loss = 'P'",
            (body.account_id, body.symbol),
        ).fetchall()
        if not lots:
            raise HTTPException(status_code=404, detail="No open positions found")

        lots = [dict(r) for r in lots]
        total_vol = sum(float(l["volume"]) for l in lots)
        avg_cost  = sum(float(l["price_entry"]) * float(l["volume"]) for l in lots) / total_vol
        exit_price = float(body.sell_price)
        closed_ids = []

        for pos in lots:
            vol = float(pos["volume"])
            pnl = round((exit_price - avg_cost) * vol, 2)
            pnl_pct = round(((exit_price / avg_cost) - 1) * 100, 2) if avg_cost > 0 else 0
            pnl_net = round(pnl - float(body.commission) / len(lots), 2)
            wl = "W" if pnl_net >= 0 else "L"
            exit_fx = _capture_thb_rate(
                _position_currency(pos), body.sell_date, conn=conn
            )
            conn.execute(
                """UPDATE trades SET date_exit = ?, price_exit = ?,
                   pnl_amount = ?, win_loss = ?, pnl_percent = ?,
                   exit_exchange_rate = ?, note = note || ?
                   WHERE id = ?""",
                (body.sell_date, exit_price, pnl_net, wl, pnl_pct, exit_fx,
                  f"\n[SOLD {body.sell_date}] @ {exit_price} | P&L: {pnl_net}",
                  pos["id"]),
            )
            closed_ids.append(pos["id"])

        return {"ok": True, "action": "sell_all_lots", "closed_ids": closed_ids, "lots_closed": len(closed_ids)}


# ── Cash Ledger ───────────────────────────────────────────────────────────────

@router.get("/cash")
def list_cash(account_id: Optional[str] = Query(None)):
    where, params = [], []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)
    sql = "SELECT * FROM cash_ledger"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY date DESC"
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


@router.post("/cash", status_code=201)
def add_cash(body: CashIn):
    entry_id = str(uuid.uuid4())
    with get_db() as conn:
        conn.execute("""
            INSERT INTO cash_ledger (id, account_id, date, income, investment, exchange_rate, note)
            VALUES (?,?,?,?,?,?,?)
        """, (entry_id, body.account_id, body.date, body.income, body.investment,
              body.exchange_rate, body.note))
    return {"ok": True, "id": entry_id}


@router.post("/cash/transfer", status_code=201)
def transfer_cash(body: CashTransferIn):
    if body.from_account_id == body.to_account_id:
        raise HTTPException(status_code=400, detail="from/to account must differ")
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="amount must be positive")
    link_id = str(uuid.uuid4())
    out_id, in_id = str(uuid.uuid4()), str(uuid.uuid4())
    with get_db() as conn:
        conn.execute(
            """INSERT INTO cash_ledger (id, account_id, date, income, investment,
               exchange_rate, note, entry_type, linked_id)
               VALUES (?,?,?,0,?,1,?,'TRANSFER',?)""",
            (out_id, body.from_account_id, body.date, -body.amount,
             body.note or f"Transfer -> {body.to_account_id}", link_id),
        )
        conn.execute(
            """INSERT INTO cash_ledger (id, account_id, date, income, investment,
               exchange_rate, note, entry_type, linked_id)
               VALUES (?,?,?,0,?,1,?,'TRANSFER',?)""",
            (in_id, body.to_account_id, body.date, body.amount,
             body.note or f"Transfer <- {body.from_account_id}", link_id),
        )
    return {"ok": True, "linked_id": link_id, "ids": [out_id, in_id]}


@router.put("/cash/{entry_id}")
def update_cash(entry_id: str, body: CashIn):
    with get_db() as conn:
        cur = conn.execute("""
            UPDATE cash_ledger SET account_id=?, date=?, income=?, investment=?,
                exchange_rate=?, note=?
            WHERE id = ?
        """, (body.account_id, body.date, body.income, body.investment,
              body.exchange_rate, body.note, entry_id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cash entry not found")
    return {"ok": True}


@router.delete("/cash/{entry_id}")
def delete_cash(entry_id: str):
    with get_db() as conn:
        row = conn.execute(
            "SELECT entry_type, linked_id FROM cash_ledger WHERE id = ?", (entry_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cash entry not found")
        if row["entry_type"] == "TRANSFER" and row["linked_id"]:
            conn.execute("DELETE FROM cash_ledger WHERE linked_id = ?", (row["linked_id"],))
        else:
            conn.execute("DELETE FROM cash_ledger WHERE id = ?", (entry_id,))
    return {"ok": True}


# ── Dividends ─────────────────────────────────────────────────────────────────

@router.get("/dividends")
def list_dividends(
    account_id: Optional[str] = Query(None),
    base_currency: Optional[str] = Query(None),
):
    where, params = [], []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)
    sql = "SELECT * FROM dividends"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY pay_date DESC"
    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
        dividends = [dict(r) for r in rows]
        if base_currency:
            base = report_currency(base_currency)
            for dividend in dividends:
                date = dividend.get("pay_date") or dividend.get("ex_date")
                ccy = dividend.get("currency") or "THB"
                dividend["amount_per_unit_base"] = round(
                    convert_amount(
                        _to_float_or_zero(dividend.get("amount_per_unit")),
                        ccy,
                        base,
                        date=date,
                        conn=conn,
                    ),
                    6,
                )
                dividend["total_received_base"] = round(
                    convert_amount(
                        _to_float_or_zero(dividend.get("total_received")),
                        ccy,
                        base,
                        date=date,
                        conn=conn,
                    ),
                    2,
                )
                dividend["reinvested_amount_base"] = round(
                    convert_amount(
                        _to_float_or_zero(dividend.get("reinvested_amount")),
                        ccy,
                        base,
                        date=date,
                        conn=conn,
                    ),
                    2,
                )
    return dividends


@router.post("/dividends", status_code=201)
def add_dividend(body: DividendIn):
    div_id = str(uuid.uuid4())
    with get_db() as conn:
        ccy_row = conn.execute(
            """
            SELECT currency, market, resolved_symbol, symbol
            FROM trades WHERE account_id = ? AND upper(symbol) = upper(?)
            ORDER BY date_entry DESC LIMIT 1
            """,
            (body.account_id, body.asset),
        ).fetchone()
        acc = conn.execute(
            "SELECT currency FROM portfolio_accounts WHERE id = ?", (body.account_id,)
        ).fetchone()
        if not acc:
            raise HTTPException(status_code=404, detail="Unknown account")
        # A matching trade is authoritative. This also protects manual forms
        # whose blank/default currency may still reflect the account rather
        # than the asset selected by the user.
        currency = None
        if ccy_row:
            currency = _position_currency(
                {**dict(ccy_row), "acc_currency": acc["currency"]}
            )
        currency = (
            currency
            or normalize_currency(body.currency)
            or normalize_currency(acc["currency"], "THB")
            or "THB"
        )
        conn.execute("""
            INSERT INTO dividends (id, account_id, asset, ex_date, pay_date,
                amount_per_unit, total_received, reinvested_amount,
                reinvest_asset, reinvest_price, reinvest_units, currency)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (div_id, body.account_id, body.asset, body.ex_date, body.pay_date,
              body.amount_per_unit, body.total_received, body.reinvested_amount,
              body.reinvest_asset, body.reinvest_price, body.reinvest_units, currency))
    return {"ok": True, "id": div_id, "currency": currency}


@router.put("/dividends/{div_id}")
def update_dividend(div_id: str, body: DividendIn):
    with get_db() as conn:
        old = conn.execute("SELECT currency FROM dividends WHERE id = ?", (div_id,)).fetchone()
        if not old:
            raise HTTPException(status_code=404, detail="Dividend entry not found")
        trade = conn.execute(
            """SELECT currency, market, resolved_symbol, symbol FROM trades
               WHERE account_id = ? AND upper(symbol) = upper(?)
               ORDER BY date_entry DESC LIMIT 1""",
            (body.account_id, body.asset),
        ).fetchone()
        acc = conn.execute(
            "SELECT currency FROM portfolio_accounts WHERE id = ?", (body.account_id,)
        ).fetchone()
        currency = None
        if trade:
            currency = _position_currency(
                {**dict(trade or {}), "acc_currency": acc["currency"] if acc else "THB"}
            )
        currency = (
            currency
            or normalize_currency(body.currency)
            or normalize_currency(old["currency"])
            or normalize_currency(acc["currency"] if acc else None, "THB")
            or "THB"
        )
        cur = conn.execute("""
            UPDATE dividends SET account_id=?, asset=?, ex_date=?, pay_date=?,
                amount_per_unit=?, total_received=?, reinvested_amount=?,
                reinvest_asset=?, reinvest_price=?, reinvest_units=?, currency=?
            WHERE id = ?
        """, (body.account_id, body.asset, body.ex_date, body.pay_date,
              body.amount_per_unit, body.total_received, body.reinvested_amount,
              body.reinvest_asset, body.reinvest_price, body.reinvest_units,
              currency, div_id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Dividend entry not found")
    return {"ok": True}


@router.delete("/dividends/{div_id}")
def delete_dividend(div_id: str):
    with get_db() as conn:
        cur = conn.execute("DELETE FROM dividends WHERE id = ?", (div_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Dividend entry not found")
    return {"ok": True}


# ── Dividend Suggestions (from yfinance) ─────────────────────────────────────

@router.get("/dividend-suggestions")
def get_dividend_suggestions(account_id: Optional[str] = Query(None)):
    """Fetch dividend data from yfinance for open positions, from the date each
    position was entered onward — never dividends predating the purchase.

    Returns one suggestion per symbol with: asset, amount_per_unit, ex_date, pay_date.
    Skips ex-dates already recorded in the dividends table (no duplicate suggestions).
    """
    where = ["win_loss = 'P'"]
    params = []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)

    with get_db() as conn:
        rows = conn.execute(
            "SELECT symbol, account_id, MIN(date_entry) date_entry, "
            "MAX(resolved_symbol) resolved_symbol, MAX(market) market, "
            "MAX(currency) currency FROM trades t "
            f"WHERE {' AND '.join(where)} GROUP BY symbol, account_id ORDER BY symbol",
            params,
        ).fetchall()

        existing = {
            (r["asset"], r["account_id"], r["ex_date"])
            for r in conn.execute("SELECT asset, account_id, ex_date FROM dividends").fetchall()
        }

    if not rows:
        return {"suggestions": []}

    suggestions = []
    for r in rows:
        symbol = r["symbol"]
        acct = r["account_id"]
        held_since = r["date_entry"]
        try:
            yf_sym = _trade_yf_symbol(dict(r))
            if not yf_sym:
                continue
            # Typed dividend access via market_data contract
            div_records = market_data.get_dividends(yf_sym)
            if not div_records:
                continue
            # Only dividends paid while the position was actually held
            eligible = [d for d in div_records if d.ex_date >= held_since]
            eligible = [d for d in eligible if (symbol, acct, d.ex_date) not in existing]
            if not eligible:
                continue
            for d in sorted(eligible, key=lambda d: d.ex_date):
                if d.amount > 0:
                    suggestions.append({
                        "asset": symbol,
                        "account_id": acct,
                        "amount_per_unit": round(d.amount, 6),
                        "ex_date": d.ex_date,
                        "pay_date": d.ex_date,  # ex-date used for pay_date
                        "currency": _position_currency(dict(r)),
                        "source": "yfinance",
                    })
        except Exception:
            continue

    return {"suggestions": suggestions}


# ── Open Positions (with live prices) ────────────────────────────────────────

@router.get("/open-positions")
def get_open_positions(
    account_id: Optional[str] = Query(None),
    base_currency: str = Query("THB"),
):
    base_currency = report_currency(base_currency)
    where = ["win_loss = 'P'"]
    params = []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)

    with get_db() as conn:
        rows = conn.execute(
            "SELECT t.*, a.currency acc_currency, a.name acc_name "
            "FROM trades t JOIN portfolio_accounts a ON t.account_id = a.id "
            f"WHERE {' AND '.join(where)} ORDER BY t.date_entry DESC",
            params,
        ).fetchall()

    positions = [dict(r) for r in rows]
    if not positions:
        return {"positions": [], "thb_per_usd": _get_thb_per_usd()}

    thb_per_usd = _get_thb_per_usd()

    # ── Batch-fetch all prices in one yfinance call ───────────────────────────
    yf_sym_map: dict[str, str] = {}   # yf_symbol → original position symbol+acct key
    pos_yf_map: dict[int, str] = {}   # position index → yf_symbol

    for i, pos in enumerate(positions):
        yf_sym = _trade_yf_symbol(pos)
        if yf_sym:
            yf_sym_map[yf_sym] = yf_sym
            pos_yf_map[i] = yf_sym

    price_lookup = _batch_fetch_prices(list(yf_sym_map.keys()))

    # ── Enrich positions with prices ─────────────────────────────────────────
    enriched = []
    for i, pos in enumerate(positions):
        yf_sym = pos_yf_map.get(i)
        quote   = price_lookup.get(yf_sym) if yf_sym else {}
        price   = quote.get("price") if quote else None
        prev_close = quote.get("prev_close") if quote else None
        pos["current_price"] = price
        pos["prev_close"] = prev_close
        # Instrument currency, not account currency — a .BK position inside a
        # USD account is priced in THB and must NOT be scaled by thb_per_usd.
        pos_ccy = _position_currency(pos)
        pos["pos_currency"] = pos_ccy
        pos["currency"] = pos_ccy
        if price and pos["price_entry"] and pos["price_entry"] > 0:
            vol = _to_float_or_zero(pos["volume"])
            entry = _to_float_or_zero(pos["price_entry"])
            pnl = (price - entry) * vol
            pos["unrealized_pnl"] = round(pnl, 4)
            pos["unrealized_pct"] = round((price - entry) / entry * 100, 2)
            pos["unrealized_pnl_thb"] = round(
                convert_amount(pnl, pos_ccy, "THB", stored_thb_rate=thb_per_usd), 2
            )
            pos["unrealized_pnl_base"] = round(
                convert_amount(pnl, pos_ccy, base_currency, stored_thb_rate=thb_per_usd), 2
            )
            cost_native = (_to_float_or_zero(pos.get("amount")) or entry * vol)
            pos["cost_basis_base"] = round(
                trade_value_in_report(pos, cost_native, base_currency, when="entry"), 2
            )
            pos["market_value_base"] = round(
                trade_value_in_report(pos, price * vol, base_currency, when="live"), 2
            )
        else:
            pos["unrealized_pnl"] = None
            pos["unrealized_pct"] = None
            pos["unrealized_pnl_thb"] = None
            pos["unrealized_pnl_base"] = None
            pos["cost_basis_base"] = None
            pos["market_value_base"] = None
        # ── Day P&L (today's move vs previous close) ──────────────────────────
        if price and prev_close and prev_close > 0:
            vol = _to_float_or_zero(pos["volume"])
            day_pnl = (price - prev_close) * vol
            pos["day_pnl"] = round(day_pnl, 4)
            pos["day_pct"] = round((price - prev_close) / prev_close * 100, 2)
            pos["day_pnl_thb"] = round(
                convert_amount(day_pnl, pos_ccy, "THB", stored_thb_rate=thb_per_usd), 2
            )
            pos["day_pnl_base"] = round(
                convert_amount(day_pnl, pos_ccy, base_currency, stored_thb_rate=thb_per_usd), 2
            )
        else:
            pos["day_pnl"] = None
            pos["day_pnl_thb"] = None
            pos["day_pnl_base"] = None
            pos["day_pct"] = None
        enriched.append(pos)

    return {"positions": enriched, "thb_per_usd": thb_per_usd}


# ── Pre-/Post-Market Session Quotes ──────────────────────────────────────────
# Pre-market prices are NOT in fast_info; they live in the full `.info` dict
# (preMarketPrice / postMarketPrice / marketState). `.info` is heavier than the
# batch fast_info used by /open-positions, so this is a SEPARATE endpoint the
# frontend fetches in the background — it never blocks the positions table.

_premarket_cache: TTLCache = TTLCache(ttl=30, maxsize=300)


def _session_pct(price: Optional[float], change: Optional[float]) -> Optional[float]:
    """Percent move, derived from change/reference so it is scaling-agnostic
    (Yahoo's *ChangePercent fields are decimal in some yfinance versions,
    percent in others — deriving avoids that ambiguity)."""
    if price is None or change is None:
        return None
    ref = price - change
    return round((change / ref) * 100, 2) if ref else None


def _fetch_session_quote(yf_sym: str) -> dict:
    try:
        raw = market_data.get_info(yf_sym).raw or {}
    except Exception:
        return {}
    pre_price = _to_float(raw.get("preMarketPrice"))
    pre_change = _to_float(raw.get("preMarketChange"))
    post_price = _to_float(raw.get("postMarketPrice"))
    post_change = _to_float(raw.get("postMarketChange"))
    return {
        "market_state": raw.get("marketState"),
        "regular_price": _to_float(raw.get("regularMarketPrice")),
        "pre_price": pre_price,
        "pre_change": pre_change,
        "pre_change_pct": _session_pct(pre_price, pre_change),
        "post_price": post_price,
        "post_change": post_change,
        "post_change_pct": _session_pct(post_price, post_change),
    }


@router.get("/premarket")
def get_premarket(account_id: Optional[str] = Query(None)):
    """Pre-/post-market session quotes for open positions, keyed by bare symbol.

    marketState ∈ {PRE, REGULAR, POST, POSTPOST, CLOSED, ...}. Pre-market is a
    US-session concept — SET (.BK) rows return no pre/post fields, which the
    frontend renders as "—".
    """
    where = ["win_loss = 'P'"]
    params: list = []
    if account_id and account_id != "all":
        where.append("account_id = ?")
        params.append(account_id)
    with get_db() as conn:
        rows = conn.execute(
            f"SELECT symbol, resolved_symbol, market, account_id FROM trades "
            f"WHERE {' AND '.join(where)}",
            params,
        ).fetchall()

    # bare symbol → yf symbol (dedup: one fetch per instrument)
    sym_map: dict[str, str] = {}
    for pos in (dict(r) for r in rows):
        yf_sym = _trade_yf_symbol(pos)
        if yf_sym:
            sym_map[str(pos["symbol"]).upper()] = yf_sym

    result: dict[str, dict] = {}
    to_fetch: dict[str, str] = {}
    for bare, yf_sym in sym_map.items():
        cached = _premarket_cache.get(yf_sym)
        if cached is not None:
            result[bare] = cached
        else:
            to_fetch[bare] = yf_sym

    if to_fetch:
        with ThreadPoolExecutor(max_workers=min(8, len(to_fetch))) as ex:
            fetched = list(ex.map(_fetch_session_quote, to_fetch.values()))
        for (bare, yf_sym), quote in zip(to_fetch.items(), fetched):
            _premarket_cache.set(yf_sym, quote)
            result[bare] = quote

    return {"quotes": result}


# ── Summary ───────────────────────────────────────────────────────────────────

@router.get("/summary")
def get_summary(base_currency: str = Query("THB")):
    base_currency = report_currency(base_currency)
    with get_db() as conn:
        accounts = [dict(r) for r in conn.execute(
            "SELECT * FROM portfolio_accounts WHERE is_active = 1"
        ).fetchall()]

        year_start = f"{datetime.now().year}-01-01"
        trade_stats = conn.execute("""
            SELECT account_id,
                   COUNT(*) total_trades,
                   SUM(CASE WHEN win_loss = 'W' THEN 1 ELSE 0 END) wins,
                   SUM(CASE WHEN win_loss = 'L' THEN 1 ELSE 0 END) losses,
                   SUM(CASE WHEN win_loss = 'P' THEN 1 ELSE 0 END) open_count,
                   SUM(CASE WHEN win_loss != 'P' AND date_exit >= ?
                            THEN 1 ELSE 0 END) ytd_closed
            FROM trades
            GROUP BY account_id
        """, (year_start,)).fetchall()

        # Realized P&L must be converted per-trade by the instrument's own market
        # currency (a .BK trade is THB even inside a USD account), then summed —
        # summing first and converting by account currency mis-scales cross-market
        # positions. pnl_amount is stored in the instrument's native currency.
        closed_pnl_rows = conn.execute("""
            SELECT t.*, a.currency acc_currency
            FROM trades t JOIN portfolio_accounts a ON t.account_id = a.id
            WHERE t.win_loss != 'P'
        """).fetchall()

        cash_stats = conn.execute("""
            SELECT account_id,
                   SUM(income) total_income,
                   SUM(investment) total_invested
            FROM cash_ledger GROUP BY account_id
        """).fetchall()

        dividend_rows = conn.execute("""
            SELECT d.*, a.currency acc_currency
            FROM dividends d JOIN portfolio_accounts a ON d.account_id = a.id
        """).fetchall()

    thb_per_usd = _get_thb_per_usd()

    stats_map = {r["account_id"]: dict(r) for r in trade_stats}
    cash_map  = {r["account_id"]: dict(r) for r in cash_stats}
    div_map: dict[str, float] = {}
    div_base_map: dict[str, float] = {}
    for raw in dividend_rows:
        row = dict(raw)
        aid = row["account_id"]
        amount = convert_amount(
            float(row.get("total_received") or 0),
            row.get("currency") or row.get("acc_currency"),
            row.get("acc_currency") or "THB",
            date=row.get("pay_date") or row.get("ex_date"),
        )
        div_map[aid] = div_map.get(aid, 0.0) + amount
        base_amount = convert_amount(
            float(row.get("total_received") or 0),
            row.get("currency") or row.get("acc_currency"),
            base_currency,
            date=row.get("pay_date") or row.get("ex_date"),
        )
        div_base_map[aid] = div_base_map.get(aid, 0.0) + base_amount

    # Per-account realized P&L, converted per-trade by instrument market currency.
    # `_acct` = in the account's own currency (for the native display column);
    # `_base` = in the requested report currency (for cross-account totals).
    pnl_acct_map:  dict[str, float] = {}
    pnl_base_map:  dict[str, float] = {}
    economic_pnl_acct_map: dict[str, float] = {}
    economic_pnl_base_map: dict[str, float] = {}
    ytd_acct_map:  dict[str, float] = {}
    ytd_base_map:  dict[str, float] = {}
    ytd_economic_acct_map: dict[str, float] = {}
    ytd_economic_base_map: dict[str, float] = {}
    for r in closed_pnl_rows:
        row = dict(r)
        aid = row["account_id"]
        acct_ccy = str(row.get("acc_currency") or "USD").upper()
        in_acct = realized_pnl_in_report(row, acct_ccy)
        in_base = realized_pnl_in_report(row, base_currency)
        economic_in_acct = realized_economic_pnl_in_report(row, acct_ccy)
        economic_in_base = realized_economic_pnl_in_report(row, base_currency)
        pnl_acct_map[aid] = pnl_acct_map.get(aid, 0.0) + in_acct
        pnl_base_map[aid] = pnl_base_map.get(aid, 0.0) + in_base
        economic_pnl_acct_map[aid] = economic_pnl_acct_map.get(aid, 0.0) + economic_in_acct
        economic_pnl_base_map[aid] = economic_pnl_base_map.get(aid, 0.0) + economic_in_base
        if str(row.get("date_exit") or "") >= year_start:
            ytd_acct_map[aid] = ytd_acct_map.get(aid, 0.0) + in_acct
            ytd_base_map[aid] = ytd_base_map.get(aid, 0.0) + in_base
            ytd_economic_acct_map[aid] = ytd_economic_acct_map.get(aid, 0.0) + economic_in_acct
            ytd_economic_base_map[aid] = ytd_economic_base_map.get(aid, 0.0) + economic_in_base

    result = []
    total_pnl_base = 0.0
    total_economic_pnl_base = 0.0
    total_ytd_realized_base = 0.0
    total_ytd_economic_realized_base = 0.0
    total_wins = total_closed = 0

    for acc in accounts:
        aid = acc["id"]
        s   = stats_map.get(aid, {})
        c   = cash_map.get(aid, {})
        dividends_native = div_map.get(aid, 0.0)
        dividends_base = div_base_map.get(aid, 0.0)
        income_thb = float(c.get("total_income") or 0)
        invested_thb = float(c.get("total_invested") or 0)

        wins   = int(s.get("wins", 0) or 0)
        losses = int(s.get("losses", 0) or 0)
        closed = wins + losses
        total  = int(s.get("total_trades", 0) or 0)
        open_n = int(s.get("open_count", 0) or 0)
        win_rate = round(wins / closed * 100, 1) if closed > 0 else 0.0

        pnl_native = pnl_acct_map.get(aid, 0.0)
        ytd_realized_native = ytd_acct_map.get(aid, 0.0)
        pnl_base = pnl_base_map.get(aid, 0.0)
        ytd_realized_base = ytd_base_map.get(aid, 0.0)
        economic_pnl_native = economic_pnl_acct_map.get(aid, 0.0)
        economic_pnl_base = economic_pnl_base_map.get(aid, 0.0)
        ytd_economic_realized_native = ytd_economic_acct_map.get(aid, 0.0)
        ytd_economic_realized_base = ytd_economic_base_map.get(aid, 0.0)

        total_pnl_base += pnl_base
        total_economic_pnl_base += economic_pnl_base
        total_ytd_realized_base += ytd_realized_base
        total_ytd_economic_realized_base += ytd_economic_realized_base
        total_wins  += wins
        total_closed += closed

        result.append({
            "account":        acc,
            "total_trades":   total,
            "open_count":     open_n,
            "wins":           wins,
            "losses":         losses,
            "win_rate":       win_rate,
            "pnl_native":     round(pnl_native, 2),
            "pnl_base":       round(pnl_base, 2),
            "pnl_economic_native": round(economic_pnl_native, 2),
            "pnl_economic_base":   round(economic_pnl_base, 2),
            "ytd_realized_native": round(ytd_realized_native, 2),
            "ytd_realized_base":   round(ytd_realized_base, 2),
            "ytd_economic_realized_native": round(ytd_economic_realized_native, 2),
            "ytd_economic_realized_base":   round(ytd_economic_realized_base, 2),
            "ytd_closed":     int(s.get("ytd_closed", 0) or 0),
            "total_income":   income_thb,
            "total_income_base": round(_conv(income_thb, "THB", base_currency, thb_per_usd), 2),
            "total_invested": invested_thb,
            "total_invested_base": round(_conv(invested_thb, "THB", base_currency, thb_per_usd), 2),
            "total_dividends": round(dividends_native, 2),
            "total_dividends_base": round(dividends_base, 2),
        })

    return {
        "accounts":       result,
        "total_pnl_base": round(total_pnl_base, 2),
        "total_economic_pnl_base": round(total_economic_pnl_base, 2),
        "total_ytd_realized_base": round(total_ytd_realized_base, 2),
        "total_ytd_economic_realized_base": round(total_ytd_economic_realized_base, 2),
        "ytd_year":       datetime.now().year,
        "global_win_rate": round(total_wins / total_closed * 100, 1) if total_closed > 0 else 0,
        "base_currency":  base_currency,
        "thb_per_usd":    thb_per_usd,
    }


# ── Money-weighted returns (CAGR + XIRR, cost-based) ────────────────────────────

def _parse_date(s) -> Optional[datetime]:
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d")
    except Exception:
        return None


def _xirr(cashflows: list[tuple[datetime, float]]) -> Optional[float]:
    """Annualized money-weighted IRR from dated (date, amount) flows.
    Sign convention: outflows (buys) negative, inflows (sells/divs/mark-to-market)
    positive. Newton–Raphson with a bisection fallback. Returns None if no sign
    change (can't solve) or non-convergent.
    """
    if len(cashflows) < 2:
        return None
    amts = [cf[1] for cf in cashflows]
    if not (any(a > 0 for a in amts) and any(a < 0 for a in amts)):
        return None
    t0 = min(cf[0] for cf in cashflows)
    ys = [(cf[0] - t0).days / 365.0 for cf in cashflows]

    def npv(r: float) -> float:
        return sum(a / ((1.0 + r) ** y) for a, y in zip(amts, ys))

    def dnpv(r: float) -> float:
        return sum(-y * a / ((1.0 + r) ** (y + 1.0)) for a, y in zip(amts, ys))

    # Newton
    r = 0.1
    ok = False
    for _ in range(100):
        d = dnpv(r)
        if abs(d) < 1e-12:
            break
        rn = r - npv(r) / d
        if rn <= -0.9999:
            rn = -0.9999
        if abs(rn - r) < 1e-8:
            r, ok = rn, True
            break
        r = rn
    if ok and abs(npv(r)) < 1.0:
        return r

    # Bisection fallback on [-0.9999, 10]
    lo, hi = -0.9999, 10.0
    flo, fhi = npv(lo), npv(hi)
    if flo * fhi > 0:
        return None
    for _ in range(200):
        mid = (lo + hi) / 2.0
        fm = npv(mid)
        if abs(fm) < 1e-7:
            return mid
        if flo * fm < 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2.0


@router.get("/returns")
def get_portfolio_returns(
    account_id: Optional[str] = Query(None),
    base_currency: str = Query("THB"),
):
    """Cost-based annualized returns: CAGR (time-weighted growth of deployed cost)
    and XIRR (money-weighted, from actual dated cashflows). Per-account + total.
    All flows normalized to `base_currency`.
    """
    base_currency = report_currency(base_currency)
    where, params = [], []
    if account_id and account_id != "all":
        where.append("t.account_id = ?")
        params.append(account_id)
    wsql = ("WHERE " + " AND ".join(where)) if where else ""

    with get_db() as conn:
        trades = [dict(r) for r in conn.execute(
            "SELECT t.*, a.currency acc_currency, a.name acc_name "
            "FROM trades t JOIN portfolio_accounts a ON t.account_id = a.id "
            f"{wsql} ORDER BY t.date_entry ASC",
            params,
        ).fetchall()]

        dwhere, dparams = [], []
        if account_id and account_id != "all":
            dwhere.append("d.account_id = ?")
            dparams.append(account_id)
        dsql = ("WHERE " + " AND ".join(dwhere)) if dwhere else ""
        divs = [dict(r) for r in conn.execute(
            "SELECT d.account_id, d.pay_date, d.ex_date, d.total_received, "
            "d.currency, a.currency acc_currency "
            "FROM dividends d JOIN portfolio_accounts a ON d.account_id = a.id "
            f"{dsql}",
            dparams,
        ).fetchall()]

    thb_per_usd = _get_thb_per_usd()
    now = datetime.now()

    # Batch-fetch current prices for open positions
    open_trades = [t for t in trades if (t.get("win_loss") or "P") == "P"]
    idx_yf: dict[int, str] = {}
    for i, t in enumerate(open_trades):
        s = _trade_yf_symbol(t)
        if s:
            idx_yf[i] = s
    price_lookup = _batch_fetch_prices(list({s for s in idx_yf.values()}))
    open_price: dict[int, Optional[float]] = {}
    for i, t in enumerate(open_trades):
        yf_sym = idx_yf.get(i)
        q = price_lookup.get(yf_sym) if yf_sym else None
        open_price[id(t)] = (q.get("price") if q else None)

    # Per-account accumulators
    class _Acc:
        __slots__ = ("cf", "invested", "realized", "unrealized", "div",
                     "first", "name")

        def __init__(self):
            self.cf: list[tuple[datetime, float]] = []
            self.invested = 0.0
            self.realized = 0.0
            self.unrealized = 0.0
            self.div = 0.0
            self.first: Optional[datetime] = None
            self.name = ""

    accs: dict[str, _Acc] = {}

    def _acc(aid: str, name: str) -> _Acc:
        a = accs.get(aid)
        if a is None:
            a = _Acc()
            a.name = name
            accs[aid] = a
        return a

    for t in trades:
        aid = t["account_id"]
        a = _acc(aid, t.get("acc_name") or aid)
        # Instrument's market currency (cross-market safe), not account currency.
        vol = _to_float_or_zero(t.get("volume"))
        entry = _to_float_or_zero(t.get("price_entry"))
        cost_native = _to_float_or_zero(t.get("amount")) or (entry * vol)
        cost = trade_value_in_report(t, cost_native, base_currency, when="entry")
        d_entry = _parse_date(t.get("date_entry"))
        if cost <= 0 or d_entry is None:
            continue
        a.invested += cost
        a.first = d_entry if a.first is None else min(a.first, d_entry)
        a.cf.append((d_entry, -cost))

        if (t.get("win_loss") or "P") != "P":   # closed
            pnl = realized_pnl_in_report(t, base_currency)
            a.realized += pnl
            d_exit = _parse_date(t.get("date_exit")) or d_entry
            a.cf.append((d_exit, cost + pnl))
        else:                                    # open → mark to market at now
            px = open_price.get(id(t))
            mv_native = (px * vol) if (px and vol) else cost_native
            mv = trade_value_in_report(t, mv_native, base_currency, when="live")
            if px and vol:
                a.unrealized += mv - cost
            a.cf.append((now, mv))

    for d in divs:
        aid = d["account_id"]
        if aid not in accs:
            continue
        a = accs[aid]
        amt = convert_amount(
            _to_float_or_zero(d.get("total_received")),
            d.get("currency") or d.get("acc_currency") or base_currency,
            base_currency,
            date=d.get("pay_date") or d.get("ex_date"),
        )
        if amt == 0:
            continue
        a.div += amt
        pd = _parse_date(d.get("pay_date")) or now
        a.cf.append((pd, amt))

    def _metrics(a: _Acc) -> dict:
        invested = a.invested
        end_value = invested + a.realized + a.unrealized + a.div
        span_days = (now - a.first).days if a.first else 0
        cagr = None
        if invested > 0 and end_value > 0 and span_days >= 1:
            cagr = (end_value / invested) ** (365.0 / span_days) - 1.0
        xirr = _xirr(a.cf)
        return {
            "cagr_pct": round(cagr * 100, 2) if cagr is not None else None,
            "xirr_pct": round(xirr * 100, 2) if xirr is not None else None,
            "simple_pct": round((end_value - invested) / invested * 100, 2) if invested > 0 else None,
            "invested": round(invested, 2),
            "end_value": round(end_value, 2),
            "realized": round(a.realized, 2),
            "unrealized": round(a.unrealized, 2),
            "dividends": round(a.div, 2),
            "holding_days": span_days,
            "first_date": a.first.strftime("%Y-%m-%d") if a.first else None,
        }

    # Total = merge all accounts' cashflows + aggregates
    total = _Acc()
    for a in accs.values():
        total.cf.extend(a.cf)
        total.invested += a.invested
        total.realized += a.realized
        total.unrealized += a.unrealized
        total.div += a.div
        if a.first and (total.first is None or a.first < total.first):
            total.first = a.first

    return {
        "base_currency": base_currency,
        "thb_per_usd": thb_per_usd,
        "total": _metrics(total),
        "accounts": {aid: {**_metrics(a), "name": a.name} for aid, a in accs.items()},
    }


# ── NAV snapshots (daily mark-to-market history) ────────────────────────────────

def _maybe_capture_nav() -> None:
    """Capture today's portfolio NAV once per day (capture-on-view).

    Writes one row per active account plus a global 'all' row, all in THB base.
    Cheap no-op if today's snapshot already exists. Never raises into callers.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        with get_db() as conn:
            if conn.execute(
                "SELECT 1 FROM portfolio_nav_snapshots WHERE snapshot_date = ? LIMIT 1",
                (today,),
            ).fetchone():
                return

            accounts = {
                a["id"]: dict(a)
                for a in conn.execute(
                    "SELECT id, currency FROM portfolio_accounts WHERE is_active = 1"
                ).fetchall()
            }
            if not accounts:
                return

            closed_rows = [
                dict(r) for r in conn.execute(
                    """SELECT t.*, a.currency acc_currency
                       FROM trades t JOIN portfolio_accounts a ON a.id = t.account_id
                       WHERE t.win_loss != 'P'"""
                ).fetchall()
            ]
            invested_map = {
                r["account_id"]: float(r["invested"] or 0)
                for r in conn.execute(
                    "SELECT account_id, SUM(COALESCE(investment,0)) invested "
                    "FROM cash_ledger GROUP BY account_id"
                ).fetchall()
            }
            dividend_rows = [
                dict(r) for r in conn.execute(
                    """SELECT d.*, a.currency acc_currency
                       FROM dividends d JOIN portfolio_accounts a ON a.id = d.account_id"""
                ).fetchall()
            ]
            open_rows = [
                dict(r)
                for r in conn.execute(
                    """SELECT t.*, a.currency acc_currency
                       FROM trades t JOIN portfolio_accounts a ON a.id = t.account_id
                       WHERE t.win_loss = 'P'"""
                ).fetchall()
            ]

        # Live prices for all open symbols, in one batch
        yf_map: dict[int, Optional[str]] = {}
        wanted: set[str] = set()
        for i, p in enumerate(open_rows):
            ys = _trade_yf_symbol(p)
            yf_map[i] = ys
            if ys:
                wanted.add(ys)
        price_lookup = _batch_fetch_prices(list(wanted))

        from collections import defaultdict as _dd
        realized_map: dict = _dd(float)
        for row in closed_rows:
            realized_map[row["account_id"]] += realized_pnl_in_report(row, "THB")
        div_map: dict = _dd(float)
        for row in dividend_rows:
            div_map[row["account_id"]] += convert_amount(
                _to_float_or_zero(row.get("total_received")),
                row.get("currency") or row.get("acc_currency"),
                "THB",
                date=row.get("pay_date") or row.get("ex_date"),
            )
        unreal: dict = _dd(float)
        cost: dict = _dd(float)
        for i, p in enumerate(open_rows):
            acc = accounts.get(p["account_id"])
            if not acc:
                continue
            vol = _to_float_or_zero(p["volume"])
            entry = _to_float_or_zero(p["price_entry"])
            cost_native = _to_float_or_zero(p["amount"]) or entry * vol
            cost_thb = trade_value_in_report(p, cost_native, "THB", when="entry")
            cost[p["account_id"]] += cost_thb
            quote = price_lookup.get(yf_map.get(i)) if yf_map.get(i) else None
            price = quote.get("price") if quote else None
            if price and entry > 0:
                market_thb = trade_value_in_report(p, price * vol, "THB", when="live")
                unreal[p["account_id"]] += market_thb - cost_thb

        rows: list[tuple] = []
        g = {"total": 0.0, "cost": 0.0, "unreal": 0.0, "real": 0.0, "inv": 0.0, "div": 0.0}
        for aid, acc in accounts.items():
            realized = realized_map.get(aid, 0.0)
            invested = invested_map.get(aid, 0.0)          # already THB
            dividends = div_map.get(aid, 0.0)
            c = cost.get(aid, 0.0)
            u = unreal.get(aid, 0.0)
            # NAV = mark-to-market value of current holdings. Deposit-independent
            # (cash_ledger deposits are often incomplete); realized/invested/dividends
            # are stored alongside for context but do not inflate the asset value.
            total = c + u
            rows.append((aid, today, total, c, u, realized, invested, dividends))
            g["total"] += total; g["cost"] += c; g["unreal"] += u
            g["real"] += realized; g["inv"] += invested; g["div"] += dividends
        rows.append(("all", today, g["total"], g["cost"], g["unreal"],
                     g["real"], g["inv"], g["div"]))

        with get_db() as conn:
            conn.executemany("""
                INSERT INTO portfolio_nav_snapshots
                    (account_id, snapshot_date, total_value, open_cost_basis,
                     unrealized_pnl, realized_pnl, invested_capital, dividends)
                VALUES (?,?,?,?,?,?,?,?)
                ON CONFLICT(account_id, snapshot_date) DO UPDATE SET
                    total_value      = excluded.total_value,
                    open_cost_basis  = excluded.open_cost_basis,
                    unrealized_pnl   = excluded.unrealized_pnl,
                    realized_pnl     = excluded.realized_pnl,
                    invested_capital = excluded.invested_capital,
                    dividends        = excluded.dividends
            """, rows)
    except Exception:
        logger.exception("NAV snapshot capture failed")


@router.get("/nav-history")
def get_nav_history(account_id: Optional[str] = Query(None), days: int = Query(365)):
    """Daily NAV time series (THB base) for charting total asset value."""
    aid = account_id if (account_id and account_id != "all") else "all"
    with get_db() as conn:
        rows = conn.execute(
            """SELECT snapshot_date, total_value, open_cost_basis, unrealized_pnl,
                      realized_pnl, invested_capital, dividends
               FROM portfolio_nav_snapshots
               WHERE account_id = ?
               ORDER BY snapshot_date DESC LIMIT ?""",
            (aid, max(1, days)),
        ).fetchall()
    return [dict(r) for r in reversed(rows)]


# ── Analytics ─────────────────────────────────────────────────────────────────

@router.get("/analytics")
def get_analytics(account_id: Optional[str] = Query(None), base_currency: str = Query("THB")):
    # NAV snapshot is capture-on-view but must not block the first viewer of the
    # day — it marks every account to market (many yfinance calls).
    threading.Thread(target=_maybe_capture_nav, daemon=True).start()
    base_currency = report_currency(base_currency)
    where, params = ["t.win_loss != 'P'"], []
    if account_id and account_id != "all":
        where.append("t.account_id = ?")
        params.append(account_id)
    open_where = ["t.win_loss = 'P'"]
    open_params: list = []
    if account_id and account_id != "all":
        open_where.append("t.account_id = ?")
        open_params.append(account_id)
    with get_db() as conn:
        closed_rows = [dict(r) for r in conn.execute(
            "SELECT t.*, COALESCE(a.name, t.account_id) acc_name, "
            "a.currency acc_currency FROM trades t "
            "LEFT JOIN portfolio_accounts a ON t.account_id = a.id "
            f"WHERE {' AND '.join(where)}",
            params,
        ).fetchall()]
        open_rows = [dict(r) for r in conn.execute(
            "SELECT t.*, a.currency acc_currency FROM trades t "
            "LEFT JOIN portfolio_accounts a ON t.account_id = a.id "
            f"WHERE {' AND '.join(open_where)}",
            open_params,
        ).fetchall()]

    from collections import defaultdict as _dd

    def _aggregate(field: str, output_key: str, *, allow_empty: bool = False) -> list[dict]:
        buckets: dict = _dd(lambda: {"cnt": 0, "wins": 0, "pnl": 0.0, "economic_pnl": 0.0})
        for row in closed_rows:
            if field == "month":
                key = str(row.get("date_exit") or row.get("date_entry") or "")[:7]
            else:
                key = str(row.get(field) or "").strip()
            if not key and not allow_empty:
                continue
            agg = buckets[key or "Other"]
            agg["cnt"] += 1
            agg["wins"] += 1 if row.get("win_loss") == "W" else 0
            agg["pnl"] += realized_pnl_in_report(row, base_currency)
            agg["economic_pnl"] += realized_economic_pnl_in_report(row, base_currency)
        result = [
            {
                output_key: key,
                "cnt": val["cnt"],
                "wins": val["wins"],
                "pnl": round(val["pnl"], 2),
                "economic_pnl": round(val["economic_pnl"], 2),
            }
            for key, val in buckets.items()
        ]
        if field == "month":
            return sorted(result, key=lambda item: item[output_key])
        return sorted(result, key=lambda item: -item["pnl"])

    by_sector = _aggregate("sector", "sector")
    by_strategy = _aggregate("strategy_name", "strategy_name")
    by_month = _aggregate("month", "month")
    top_symbols = _aggregate("symbol", "symbol")[:15]
    subport_rows = [
        {
            "note": row.get("note"),
            "win_loss": row.get("win_loss"),
            "acc_name": row.get("acc_name") or row.get("account_id"),
            "pnl": realized_pnl_in_report(row, base_currency),
        }
        for row in closed_rows
    ]

    open_sector_totals: dict[str, float] = _dd(float)
    open_symbols: list[dict] = []
    for row in open_rows:
        sector = str(row.get("sector") or "Other")
        cost_native = _to_float_or_zero(row.get("amount")) or (
            _to_float_or_zero(row.get("price_entry")) * _to_float_or_zero(row.get("volume"))
        )
        cost_base = trade_value_in_report(row, cost_native, base_currency, when="entry")
        open_sector_totals[sector] += cost_base
        open_symbols.append({
            "sector": sector,
            "symbol": row.get("symbol"),
            "volume": row.get("volume"),
            "cost_base": cost_base,
        })
    open_by_sector = [
        {"sector": key, "cost_base": value}
        for key, value in sorted(open_sector_totals.items(), key=lambda item: -item[1])
    ]

    # group symbols by sector
    syms_by_sector: dict = _dd(list)
    for r in open_symbols:
        cost = round(r["cost_base"] or 0, 0)
        syms_by_sector[r["sector"]].append({
            "symbol": r["symbol"],
            "volume": r["volume"],
            "value":  cost,
        })

    # Sub-port breakdown. New trades store "AccName (subport-id) | freeform | VAT: x"
    # (see helpers.ts splitNote). Sold trades additionally get "\n[SOLD ...]"
    # appended directly (see /sell, /sell-all-lots) — so the sub-port tag is only
    # ever guaranteed to be the leading segment of the note. Match from the start.
    def _extract_subport(note: Optional[str]) -> Optional[str]:
        if not note:
            return None
        m = re.match(r"^[^(\n|]*\(([^)]+)\)", note.strip())
        return m.group(1) if m else None

    subport_agg: dict = _dd(lambda: {"cnt": 0, "wins": 0, "pnl": 0.0})
    for r in subport_rows:
        sub = _extract_subport(r["note"])
        label = f"{r['acc_name']} ({sub})" if sub else r["acc_name"]
        agg = subport_agg[label]
        agg["cnt"] += 1
        agg["wins"] += 1 if r["win_loss"] == "W" else 0
        agg["pnl"] += r["pnl"] or 0
    by_subport = sorted(
        (
            {"subport": k, "cnt": v["cnt"], "wins": v["wins"], "pnl": round(v["pnl"], 2)}
            for k, v in subport_agg.items()
        ),
        key=lambda d: -d["pnl"],
    )

    def _fmt(rows) -> list[dict]:
        result = []
        for r in rows:
            d = dict(r)
            cnt = d.get("cnt", 1) or 1
            d["win_rate"] = round(d.get("wins", 0) / cnt * 100, 1)
            result.append(d)
        return result

    return {
        "by_sector":     _fmt(by_sector),
        "by_strategy":   _fmt(by_strategy),
        "by_month":      _fmt(by_month),
        "top_symbols":   _fmt(top_symbols),
        "by_subport":    _fmt(by_subport),
        "open_by_sector": [
            {
                "sector":  r["sector"],
                "value":   round(r["cost_base"] or 0, 2),
                "symbols": syms_by_sector[r["sector"]],
            }
            for r in open_by_sector if (r["cost_base"] or 0) > 0
        ],
    }


# ── Allocation History ────────────────────────────────────────────────────────

@router.get("/analytics/allocation-history")
def get_allocation_history(
    period: str = Query("Q"),
    account_id: Optional[str] = Query(None),
):
    """
    Return cost basis (THB) deployed per sector per time period.
    Groups by date_entry so each period shows capital actually invested that period.
    period: "M" = monthly (YYYY-MM), "Q" = quarterly (YYYY-QN)
    """
    if period not in ("M", "Q"):
        raise HTTPException(status_code=400, detail="period must be M or Q")

    where_parts: list[str] = []
    params: list = []
    if account_id and account_id != "all":
        where_parts.append("t.account_id = ?")
        params.append(account_id)
    with get_db() as conn:
        sql = """SELECT t.*, a.currency acc_currency FROM trades t
                 JOIN portfolio_accounts a ON t.account_id = a.id
                 WHERE t.date_entry IS NOT NULL AND t.date_entry != ''"""
        if where_parts:
            sql += " AND " + " AND ".join(where_parts)
        rows = [dict(r) for r in conn.execute(sql, params).fetchall()]

    # pivot: {period → {sector → value}}
    from collections import defaultdict
    data: dict[str, dict[str, float]] = defaultdict(dict)
    sectors: set[str] = set()
    for row in rows:
        date = str(row.get("date_entry") or "")[:10]
        if period == "M":
            period_key = date[:7]
        else:
            month = int(date[5:7])
            period_key = f"{date[:4]}-Q{((month - 1) // 3) + 1}"
        sector = str(row.get("sector") or "Other")
        cost_native = _to_float_or_zero(row.get("amount")) or (
            _to_float_or_zero(row.get("price_entry")) * _to_float_or_zero(row.get("volume"))
        )
        v = trade_value_in_report(row, cost_native, "THB", when="entry")
        if v > 0:
            data[period_key][sector] = round(data[period_key].get(sector, 0) + v, 0)
            sectors.add(sector)

    # normalize — every period must have every sector key (0 if absent)
    # recharts stacked area requires consistent keys across all data points
    all_sectors = sorted(sectors)
    periods_out = [
        {"period": p, **{s: data[p].get(s, 0) for s in all_sectors}}
        for p in sorted(data.keys())
    ]
    return {"periods": periods_out, "sectors": all_sectors}


# ── Import from Excel ─────────────────────────────────────────────────────────

@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """
    Parse xlsx file with sheets: Finansia, Dime, InnovestX, Income&expenses.
    Idempotent: uses a stable ID based on account+symbol+date_entry+volume.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    # Map sheet name → (account_id, currency)
    sheet_map = {
        "Finansia":  ("finansia",  "THB"),
        "Finansia ": ("finansia",  "THB"),
        "Dime":      ("dime",      "USD"),
        "InnovestX": ("innovestx", "THB"),
        "InnovestX ":("innovestx", "THB"),
    }

    all_trades: list[dict] = []
    all_cash:   list[dict] = []
    all_divs:   list[dict] = []
    now = _now()

    # ── Parse trade sheets ────────────────────────────────────────────────────
    for sheet_name, (account_id, currency) in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_idx: dict[str, int] = {}

        for row_vals in ws.iter_rows(values_only=True):
            # Skip empty rows
            if not any(v is not None for v in row_vals):
                continue

            # Detect header row
            if not header_idx:
                row0 = str(row_vals[0] or "").strip().lower()
                if "date" in row0:
                    for i, h in enumerate(row_vals):
                        if h is not None:
                            header_idx[str(h).strip().lower()] = i
                continue

            def _get(*keys):
                for k in keys:
                    for hk, hi in header_idx.items():
                        if k in hk and hi < len(row_vals):
                            v = row_vals[hi]
                            if v is not None:
                                return v
                return None

            symbol = _get("stock")
            if not symbol:
                continue
            date_entry = _fmt_date(_get("date (entry)"))
            if not date_entry:
                continue

            vol = _to_float_or_zero(_get("volume"))
            price_e = _to_float_or_zero(_get("price_entry"))

            # Stable ID: hash of key fields to avoid duplicate imports
            stable_key = f"{account_id}|{str(symbol).upper()}|{date_entry}|{vol:.4f}|{price_e:.4f}"
            trade_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, stable_key))

            win_loss_raw = str(_get("win/loss") or "P").strip().upper()
            win_loss = win_loss_raw if win_loss_raw in ("W", "L", "P") else "P"

            all_trades.append({
                "id":               trade_id,
                "account_id":       account_id,
                "symbol":           str(symbol).strip().upper(),
                "sector":           str(_get("sector") or "").strip(),
                "date_entry":       date_entry,
                "date_exit":        _fmt_date(_get("date (exit)")),
                "price_entry":      price_e,
                "price_exit":       _to_float(_get("price_exit")),
                "price_stoploss":   _to_float(_get("price_stoploss")),
                "price_target":     _to_float(_get("price_target")),
                "volume":           vol,
                "amount":           _to_float(_get("amount")),
                "pnl_amount":       _to_float(_get("pnl_amount")),
                "win_loss":         win_loss,
                "pnl_percent":      _to_float(_get("pnl_percent")),
                "currency":         infer_instrument_currency(
                    None, _get("resolved symbol"), symbol, currency
                ) or currency,
                "exchange_rate":    _to_float_or_zero(_get("exchange rate", "fx rate")) or 1.0,
                "strategy_name":    str(_get("strategy name", "strategy_name") or ""),
                "entry_trigger":    str(_get("entry trigger") or ""),
                "exit_trigger":     str(_get("exit trigger") or ""),
                "market_trend":     str(_get("market_trend") or ""),
                "news_sentiment":   str(_get("news_sentiment") or ""),
                "expectation_based":str(_get("expectation_based") or ""),
                "factor_based":     str(_get("factor_based") or ""),
                "fear_greed_index": str(_get("fear & greed index") or ""),
                "vix_index":        str(_get("vix index") or ""),
                "note":             str(_get("note") or ""),
                "created_at":       now,
            })

    # ── Parse Income&expenses sheet ───────────────────────────────────────────
    ie_names = [n for n in wb.sheetnames if "income" in n.lower() or "expense" in n.lower()]
    if ie_names:
        ws_ie = wb[ie_names[0]]
        header_found = False
        for row_vals in ws_ie.iter_rows(values_only=True):
            if not any(v is not None for v in row_vals):
                continue
            # Detect header: contains วัน/เดือน/ปี or date-like Thai text
            if not header_found:
                r0 = str(row_vals[0] or "").strip()
                if "วัน" in r0 or "date" in r0.lower():
                    header_found = True
                continue

            # Cash section: cols 0–4
            date_val    = row_vals[0] if len(row_vals) > 0 else None
            income_val  = row_vals[1] if len(row_vals) > 1 else None
            invest_val  = row_vals[2] if len(row_vals) > 2 else None
            account_val = row_vals[3] if len(row_vals) > 3 else None
            fx_val      = row_vals[4] if len(row_vals) > 4 else None

            date_str = _fmt_date(date_val)
            if date_str and (income_val is not None or invest_val is not None):
                acc_id = _map_account_name(str(account_val or ""))
                income   = _to_float_or_zero(income_val)
                invested = _to_float_or_zero(invest_val)
                if income != 0 or invested != 0:
                    stable_key = f"cash|{acc_id}|{date_str}|{income:.2f}|{invested:.2f}"
                    all_cash.append({
                        "id":           str(uuid.uuid5(uuid.NAMESPACE_DNS, stable_key)),
                        "account_id":   acc_id,
                        "date":         date_str,
                        "income":       income,
                        "investment":   invested,
                        "exchange_rate": _to_float_or_zero(fx_val) or 1.0,
                        "note":         "",
                        "created_at":   now,
                    })

            # Dividend section: cols 6–10 (ex_date, pay_date, amount/unit, net_received, account)
            ex_date_val  = row_vals[6]  if len(row_vals) > 6  else None
            pay_date_val = row_vals[7]  if len(row_vals) > 7  else None
            amt_unit_val = row_vals[8]  if len(row_vals) > 8  else None
            net_recv_val = row_vals[9]  if len(row_vals) > 9  else None
            div_acc_val  = row_vals[10] if len(row_vals) > 10 else None
            asset_val    = row_vals[5]  if len(row_vals) > 5  else None

            ex_str  = _fmt_date(ex_date_val)
            pay_str = _fmt_date(pay_date_val)
            if ex_str and net_recv_val is not None:
                net_recv = _to_float_or_zero(net_recv_val)
                if net_recv != 0:
                    div_acc = _map_account_name(str(div_acc_val or ""))
                    asset   = str(asset_val or "").strip()
                    stable_key = f"div|{div_acc}|{asset}|{ex_str}|{net_recv:.4f}"
                    all_divs.append({
                        "id":               str(uuid.uuid5(uuid.NAMESPACE_DNS, stable_key)),
                        "account_id":       div_acc,
                        "asset":            asset,
                        "ex_date":          ex_str,
                        "pay_date":         pay_str,
                        "amount_per_unit":  _to_float_or_zero(amt_unit_val),
                        "total_received":   net_recv,
                        "reinvested_amount": 0,
                        "reinvest_asset":   "",
                        "reinvest_price":   0,
                        "reinvest_units":   0,
                        "created_at":       now,
                    })

    # ── Upsert into DB ────────────────────────────────────────────────────────
    inserted = {"trades": 0, "cash": 0, "dividends": 0}
    with get_db() as conn:
        for t in all_trades:
            existing = conn.execute(
                """SELECT 1 FROM trades WHERE id = ? OR (
                       account_id = ? AND upper(symbol) = upper(?) AND date_entry = ?
                       AND ABS(volume - ?) < 0.0001 AND ABS(price_entry - ?) < 0.0001
                   )""",
                (t["id"], t["account_id"], t["symbol"], t["date_entry"], t["volume"], t["price_entry"]),
            ).fetchone()
            if not existing:
                entry_fx = _capture_thb_rate(
                    t["currency"], t["date_entry"], t["exchange_rate"], conn=conn
                )
                exit_fx = None
                if t["date_exit"] or t["win_loss"] != "P":
                    exit_fx = _capture_thb_rate(
                        t["currency"], t["date_exit"] or t["date_entry"], conn=conn
                    )
                conn.execute("""
                    INSERT INTO trades (id, account_id, symbol, sector, date_entry, date_exit,
                        price_entry, price_exit, price_stoploss, price_target, volume, amount,
                        pnl_amount, win_loss, pnl_percent, currency, exchange_rate, exit_exchange_rate,
                        strategy_name, entry_trigger, exit_trigger, market_trend,
                        news_sentiment, expectation_based, factor_based,
                        fear_greed_index, vix_index, note, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (t["id"], t["account_id"], t["symbol"], t["sector"],
                      t["date_entry"], t["date_exit"],
                      t["price_entry"], t["price_exit"], t["price_stoploss"], t["price_target"],
                      t["volume"], t["amount"], t["pnl_amount"], t["win_loss"],
                      t["pnl_percent"], t["currency"], entry_fx, exit_fx,
                      t["strategy_name"], t["entry_trigger"], t["exit_trigger"],
                      t["market_trend"], t["news_sentiment"], t["expectation_based"],
                      t["factor_based"], t["fear_greed_index"], t["vix_index"],
                      t["note"], t["created_at"]))
                inserted["trades"] += 1

        for c in all_cash:
            dup = conn.execute(
                """SELECT 1 FROM cash_ledger WHERE account_id = ? AND date = ?
                   AND ABS(income - ?) < 0.01 AND ABS(investment - ?) < 0.01""",
                (c["account_id"], c["date"], c["income"], c["investment"]),
            ).fetchone()
            if dup:
                continue
            cur = conn.execute("""
                INSERT OR IGNORE INTO cash_ledger
                    (id, account_id, date, income, investment, exchange_rate, note, created_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (c["id"], c["account_id"], c["date"], c["income"], c["investment"],
                  c["exchange_rate"], c["note"], c["created_at"]))
            inserted["cash"] += cur.rowcount

        for d in all_divs:
            dup = conn.execute(
                """SELECT 1 FROM dividends WHERE account_id = ? AND upper(asset) = upper(?)
                   AND ex_date = ? AND ABS(total_received - ?) < 0.01""",
                (d["account_id"], d["asset"], d["ex_date"], d["total_received"]),
            ).fetchone()
            if dup:
                continue
            trade = conn.execute(
                """SELECT currency, market, resolved_symbol, symbol FROM trades
                   WHERE account_id = ? AND upper(symbol) = upper(?)
                   ORDER BY date_entry DESC LIMIT 1""",
                (d["account_id"], d["asset"]),
            ).fetchone()
            account = conn.execute(
                "SELECT currency FROM portfolio_accounts WHERE id = ?", (d["account_id"],)
            ).fetchone()
            div_ccy = _position_currency({
                **dict(trade or {}),
                "acc_currency": account["currency"] if account else "THB",
            })
            cur = conn.execute("""
                INSERT OR IGNORE INTO dividends
                    (id, account_id, asset, ex_date, pay_date, amount_per_unit,
                     total_received, reinvested_amount, reinvest_asset,
                     reinvest_price, reinvest_units, currency, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (d["id"], d["account_id"], d["asset"], d["ex_date"], d["pay_date"],
                   d["amount_per_unit"], d["total_received"], d["reinvested_amount"],
                   d["reinvest_asset"], d["reinvest_price"], d["reinvest_units"],
                   div_ccy, d["created_at"]))
            inserted["dividends"] += cur.rowcount

    return {
        "ok": True,
        "inserted": inserted,
        "parsed": {"trades": len(all_trades), "cash": len(all_cash), "dividends": len(all_divs)},
    }
